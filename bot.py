#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import logging
import asyncio
import json
import re
from datetime import datetime, timedelta, time, date
from dataclasses import dataclass
from typing import Dict, Tuple, Optional, List, Any
from enum import Enum

from openpyxl import Workbook, load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ConversationHandler, filters, ContextTypes
)


# ================== НАСТРОЙКИ ==================
class Config:
    BOT_TOKEN = os.getenv("BOT_TOKEN")

    ADMIN_IDS = [6056091640]
    DATA_DIR = "data"
    TEMPLATE_FILE = "Табличка для бота по питанию.xlsx"
    ORDERS_FILE = "orders.xlsx"
    STUDENTS_FILE = "students.xlsx"
    SESSIONS_FILE = "sessions.json"
    REMINDERS_FILE = "reminders.json"
    DEADLINE_TIME = time(8, 0)  # Дедлайн - 8:00 утра
    REMINDER_TIME = time(14, 0)  # Напоминание в 7:00
    TIMEZONE_OFFSET = 2  # Смещение часового пояса (Москва UTC+3)


# Настройка логгирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

DAY_NAMES_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]


class MealType(Enum):
    BREAKFAST = "breakfast"
    LUNCH = "lunch"
    SNACK = "snack"


# ================== УТИЛИТЫ ==================
def get_current_datetime() -> datetime:
    """Получает текущее время с учетом часового пояса"""
    now = datetime.utcnow() + timedelta(hours=Config.TIMEZONE_OFFSET)
    return now


def is_date_locked(target_date: date) -> bool:
    """Проверяет, заблокирована ли дата для редактирования"""
    now = get_current_datetime()
    today = now.date()
    current_time = now.time()

    logger.info(f"Проверка блокировки: дата={target_date}, сегодня={today}, время={current_time.strftime('%H:%M:%S')}")

    # 1. Если дата уже прошла
    if target_date < today:
        logger.info(f"Дата {target_date} прошла - БЛОКИРОВАНО")
        return True

    # 2. Если сегодня и время после дедлайна (8:00)
    if target_date == today and current_time >= Config.DEADLINE_TIME:
        logger.info(
            f"Сегодня {today}, время {current_time.strftime('%H:%M')} после дедлайна {Config.DEADLINE_TIME.strftime('%H:%M')} - БЛОКИРОВАНО")
        return True

    logger.info(f"Дата {target_date} доступна для редактирования")
    return False


# ================== МОДЕЛИ ==================
@dataclass
class StudentInfo:
    student_id: str
    full_name: str
    class_name: str


# ================== МЕНЕДЖЕР НАПОМИНАНИЙ ==================
class ReminderManager:
    def __init__(self, reminders_path: str):
        self.reminders_path = reminders_path
        self.reminders = self._load_reminders()

    def _load_reminders(self) -> Dict:
        """Загружает настройки напоминаний"""
        if os.path.exists(self.reminders_path):
            try:
                with open(self.reminders_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                logger.error(f"Ошибка загрузки напоминаний: {e}")
        return {}

    def _save_reminders(self):
        """Сохраняет настройки напоминаний"""
        try:
            with open(self.reminders_path, 'w', encoding='utf-8') as f:
                json.dump(self.reminders, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Ошибка сохранения напоминаний: {e}")

    def get_user_reminder(self, user_id: int) -> bool:
        """Получает статус напоминания для пользователя"""
        return self.reminders.get(str(user_id), False)

    def set_user_reminder(self, user_id: int, enabled: bool):
        """Устанавливает статус напоминания для пользователя"""
        self.reminders[str(user_id)] = enabled
        self._save_reminders()
        logger.info(f"Напоминание для пользователя {user_id}: {'включено' if enabled else 'выключено'}")

    def get_all_users_with_reminders(self) -> List[int]:
        """Получает список всех пользователей с включенными напоминаниями"""
        return [int(user_id) for user_id, enabled in self.reminders.items() if enabled]

    def toggle_user_reminder(self, user_id: int) -> bool:
        """Переключает статус напоминания для пользователя"""
        current = self.get_user_reminder(user_id)
        new_state = not current
        self.set_user_reminder(user_id, new_state)
        return new_state


# ================== МЕНЕДЖЕР ШАБЛОНА ==================
class TemplateManager:
    def __init__(self, template_path: str):
        self.template_path = template_path
        self.workbook = None
        self.structure = {}

    def load_template(self) -> bool:
        """Загружает и анализирует шаблон"""
        if not os.path.exists(self.template_path):
            logger.error(f"Файл шаблона не найден: {self.template_path}")
            return False

        try:
            logger.info(f"Загрузка шаблона: {self.template_path}")
            self.workbook = load_workbook(self.template_path)
            self.structure = self._analyze_structure()
            logger.info(f"Шаблон загружен успешно. Листов: {len(self.workbook.sheetnames)}")
            return True
        except Exception as e:
            logger.error(f"Ошибка загрузки шаблона: {e}", exc_info=True)
            return False

    def _analyze_structure(self) -> Dict:
        """Анализирует структуру шаблона"""
        structure = {}

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            logger.info(f"Анализ листа: {sheet_name}")

            sheet_structure = {
                'class_name': sheet_name,
                'date_columns': {},  # дата -> (завтрак_кол, обед_кол, полдник_кол)
                'students': {},  # ФИО -> строка
                'date_row': None,
                'students_start_row': None
            }

            # Ищем строку с датами
            for row in range(1, 10):
                cell = sheet.cell(row=row, column=3)  # Колонка C
                if cell.value and self._is_date(cell.value):
                    sheet_structure['date_row'] = row
                    logger.info(f"Найдена строка с датами: строка {row}")
                    break

            if not sheet_structure['date_row']:
                sheet_structure['date_row'] = 3
                logger.warning(f"Не найдена строка с датами для листа {sheet_name}, используем строку 3")

            # Парсим даты
            self._parse_dates(sheet, sheet_structure)

            # Ищем начало списка учеников
            for row in range(1, 20):
                if sheet.cell(row=row, column=1).value == "пп":
                    sheet_structure['students_start_row'] = row + 1
                    logger.info(f"Начало списка учеников: строка {row + 1}")
                    break

            if not sheet_structure['students_start_row']:
                sheet_structure['students_start_row'] = 4

            # Парсим учеников
            self._parse_students(sheet, sheet_structure)

            structure[sheet_name] = sheet_structure

        return structure

    def _is_date(self, value) -> bool:
        """Проверяет, является ли значение датой"""
        if isinstance(value, datetime):
            return True

        value_str = str(value)
        date_patterns = [
            r'\d{4}-\d{2}-\d{2}',
            r'\d{2}\.\d{2}\.\d{4}',
            r'\d{2}/\d{2}/\d{4}'
        ]

        for pattern in date_patterns:
            if re.search(pattern, value_str):
                return True

        return False

    def _parse_dates(self, sheet, sheet_structure: Dict):
        """Парсит даты из шаблона"""
        date_row = sheet_structure['date_row']

        col = 3  # Начинаем с колонки C
        while col <= sheet.max_column:
            date_cell = sheet.cell(row=date_row, column=col)
            date_value = self._normalize_date(date_cell.value)

            if date_value:
                sheet_structure['date_columns'][date_value] = {
                    'breakfast_col': col,
                    'lunch_col': col + 1,
                    'snack_col': col + 2
                }
                logger.debug(f"Найдена дата {date_value} в колонках {col}-{col + 2}")
                col += 3  # Переходим к следующей дате
            else:
                col += 1

    def _parse_students(self, sheet, sheet_structure: Dict):
        """Парсит список учеников"""
        start_row = sheet_structure['students_start_row']

        for row in range(start_row, sheet.max_row + 1):
            name_cell = sheet.cell(row=row, column=2)  # Колонка B - ФИО
            if name_cell.value:
                student_name = str(name_cell.value).strip()
                if (student_name and
                        student_name != "Итого:" and
                        not student_name.startswith("Всего:")):
                    sheet_structure['students'][student_name] = row

    def _normalize_date(self, value) -> Optional[str]:
        """Приводит дату к стандартному формату YYYY-MM-DD"""
        if not value:
            return None

        try:
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")

            value_str = str(value).strip()

            # Убираем время если есть
            if " 00:00:00" in value_str:
                value_str = value_str.replace(" 00:00:00", "")

            # Пробуем разные форматы
            date_formats = [
                "%Y-%m-%d",
                "%d.%m.%Y",
                "%d/%m/%Y",
                "%d-%m-%Y"
            ]

            for fmt in date_formats:
                try:
                    dt = datetime.strptime(value_str, fmt)
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue

            # Пробуем извлечь дату из строки
            date_patterns = [
                r'\d{4}-\d{2}-\d{2}',
                r'\d{2}\.\d{2}\.\d{4}',
                r'\d{2}/\d{2}/\d{4}'
            ]

            for pattern in date_patterns:
                match = re.search(pattern, value_str)
                if match:
                    date_str = match.group()
                    for fmt in date_formats:
                        try:
                            dt = datetime.strptime(date_str, fmt)
                            return dt.strftime("%Y-%m-%d")
                        except ValueError:
                            continue

        except Exception as e:
            logger.debug(f"Ошибка нормализации даты '{value}': {e}")

        return None

    def update_order(self, student_name: str, date_str: str, meals: Dict[str, bool]) -> bool:
        """Обновляет заказ в шаблоне"""
        if not self.workbook:
            if not self.load_template():
                return False

        try:
            # Находим ученика
            sheet_name, student_row = self.find_student(student_name)
            if not sheet_name or not student_row:
                logger.error(f"Ученик не найден в шаблоне: {student_name}")
                return False

            # Находим колонки для даты
            sheet_structure = self.structure.get(sheet_name)
            if not sheet_structure:
                logger.error(f"Структура листа {sheet_name} не найдена")
                return False

            date_info = sheet_structure['date_columns'].get(date_str)
            if not date_info:
                logger.error(f"Дата {date_str} не найдена в листе {sheet_name}")
                return False

            sheet = self.workbook[sheet_name]

            # Обновляем ячейки
            if meals.get('breakfast'):
                sheet.cell(row=student_row, column=date_info['breakfast_col'], value="З")
            else:
                sheet.cell(row=student_row, column=date_info['breakfast_col'], value="")

            if meals.get('lunch'):
                sheet.cell(row=student_row, column=date_info['lunch_col'], value="О")
            else:
                sheet.cell(row=student_row, column=date_info['lunch_col'], value="")

            if meals.get('snack'):
                sheet.cell(row=student_row, column=date_info['snack_col'], value="П")
            else:
                sheet.cell(row=student_row, column=date_info['snack_col'], value="")

            # Сохраняем изменения
            self.workbook.save(self.template_path)
            logger.info(f"Шаблон обновлен: {student_name} - {date_str}")
            return True

        except Exception as e:
            logger.error(f"Ошибка обновления шаблона: {e}", exc_info=True)
            return False

    def find_student(self, student_name: str) -> Tuple[Optional[str], Optional[int]]:
        """Находит ученика в шаблоне"""
        for sheet_name, sheet_structure in self.structure.items():
            for name, row in sheet_structure['students'].items():
                if name.strip().lower() == student_name.strip().lower():
                    return sheet_name, row
        return None, None


# ================== БАЗА ДАННЫХ ==================
class Database:
    def __init__(self):
        os.makedirs(Config.DATA_DIR, exist_ok=True)
        self.template_path = os.path.join(Config.DATA_DIR, Config.TEMPLATE_FILE)
        self.orders_path = os.path.join(Config.DATA_DIR, Config.ORDERS_FILE)
        self.students_path = os.path.join(Config.DATA_DIR, Config.STUDENTS_FILE)
        self.reminders_path = os.path.join(Config.DATA_DIR, Config.REMINDERS_FILE)

        self.template_manager = TemplateManager(self.template_path)
        self.reminder_manager = ReminderManager(self.reminders_path)

        # Инициализация файлов
        self._init_files()

    def _init_files(self):
        """Инициализация всех файлов"""
        # Проверяем students.xlsx
        if not os.path.exists(self.students_path):
            logger.error(f"Файл {self.students_path} не найден!")
            return

        # Загружаем шаблон
        if os.path.exists(self.template_path):
            self.template_manager.load_template()

        # Создаем или обновляем orders.xlsx
        self._create_or_update_orders_file()

    def _create_or_update_orders_file(self):
        """Создает или обновляет файл заказов"""
        try:
            # Загружаем учеников
            student_wb = load_workbook(self.students_path, data_only=True)
            student_ws = student_wb.active

            students = []
            for row in student_ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] and row[1]:
                    students.append({
                        'id': str(row[0]),
                        'name': row[1],
                        'class': row[2] if len(row) > 2 else ""
                    })

            # Получаем даты
            dates = []
            if self.template_manager.workbook:
                # Получаем все даты из шаблона
                all_dates = []
                for sheet_structure in self.template_manager.structure.values():
                    all_dates.extend(sheet_structure['date_columns'].keys())

                # Убираем дубликаты и сортируем
                dates = sorted(list(set(all_dates)))
                logger.info(f"Загружено {len(dates)} уникальных дат из шаблона")

            if not dates:
                # Создаем даты на 30 рабочих дней вперед
                today = datetime.now()
                added = 0
                date = today
                while added < 30:
                    if date.weekday() < 5:
                        dates.append(date.strftime("%Y-%m-%d"))
                        added += 1
                    date += timedelta(days=1)

            # Проверяем существует ли orders.xlsx
            if os.path.exists(self.orders_path):
                self._update_orders_file(students, dates)
            else:
                self._create_new_orders_file(students, dates)

        except Exception as e:
            logger.error(f"Ошибка создания/обновления orders.xlsx: {e}")

    def _create_new_orders_file(self, students: List[Dict], dates: List[str]):
        """Создает новый файл заказов"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"

        # Заголовки
        headers = ["ID", "ФИО", "Класс"]
        for date_str in dates:
            headers.extend([
                f"{date_str}_breakfast",
                f"{date_str}_lunch",
                f"{date_str}_snack"
            ])

        ws.append(headers)

        # Добавляем учеников
        for student in students:
            student_row = [student['id'], student['name'], student['class']]
            student_row.extend([""] * (len(dates) * 3))
            ws.append(student_row)

        wb.save(self.orders_path)
        logger.info(f"Создан новый файл orders.xlsx")

    def _update_orders_file(self, students: List[Dict], dates: List[str]):
        """Обновляет существующий файл заказов"""
        wb = load_workbook(self.orders_path)
        ws = wb.active

        # Получаем текущие заголовки
        current_headers = []
        for col in range(1, ws.max_column + 1):
            current_headers.append(ws.cell(1, col).value)

        # Добавляем недостающие даты
        new_dates = []
        for date_str in dates:
            date_headers = [
                f"{date_str}_breakfast",
                f"{date_str}_lunch",
                f"{date_str}_snack"
            ]

            if not all(header in current_headers for header in date_headers):
                new_dates.append(date_str)

        if new_dates:
            for date_str in new_dates:
                ws.cell(1, ws.max_column + 1, f"{date_str}_breakfast")
                ws.cell(1, ws.max_column + 1, f"{date_str}_lunch")
                ws.cell(1, ws.max_column + 1, f"{date_str}_snack")

            # Добавляем пустые ячейки
            for row in range(2, ws.max_row + 1):
                for _ in range(len(new_dates) * 3):
                    ws.cell(row, ws.max_column + 1, "")

        wb.save(self.orders_path)
        if new_dates:
            logger.info(f"Добавлено {len(new_dates)} новых дат в orders.xlsx")

    def verify_student(self, student_id: str) -> Tuple[bool, Optional[StudentInfo]]:
        """Проверяет ученика по ID"""
        try:
            wb = load_workbook(self.students_path, data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == student_id:
                    return True, StudentInfo(
                        student_id=str(row[0]),
                        full_name=row[1],
                        class_name=row[2] if len(row) > 2 else ""
                    )

        except Exception as e:
            logger.error(f"Ошибка проверки ученика: {e}")

        return False, None

    def save_order(self, student_id: str, date_str: str, meals: Dict[str, bool]) -> bool:
        """Сохраняет заказ ученика"""
        try:
            # Проверяем не заблокирована ли дата
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            if is_date_locked(target_date):
                logger.warning(f"Попытка сохранить заказ на заблокированную дату: {date_str}")
                return False

            # 1. Сохраняем в orders.xlsx
            wb = load_workbook(self.orders_path)
            ws = wb.active

            # Находим строку ученика
            student_row = None
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(r, 1).value) == student_id:
                    student_row = r
                    break

            if not student_row:
                return False

            # Находим колонки для даты
            breakfast_col = None
            lunch_col = None
            snack_col = None

            for col in range(4, ws.max_column + 1):
                header = ws.cell(1, col).value
                if header and date_str in str(header):
                    if "_breakfast" in str(header):
                        breakfast_col = col
                    elif "_lunch" in str(header):
                        lunch_col = col
                    elif "_snack" in str(header):
                        snack_col = col

            if not all([breakfast_col, lunch_col, snack_col]):
                return False

            # Сохраняем заказы
            ws.cell(row=student_row, column=breakfast_col, value="✅" if meals.get('breakfast') else "")
            ws.cell(row=student_row, column=lunch_col, value="✅" if meals.get('lunch') else "")
            ws.cell(row=student_row, column=snack_col, value="✅" if meals.get('snack') else "")

            wb.save(self.orders_path)

            # 2. Обновляем шаблон
            ok, student = self.verify_student(student_id)
            if ok and student.full_name:
                self.template_manager.update_order(student.full_name, date_str, meals)

            logger.info(f"Заказ сохранен: ID {student_id} - {date_str}")
            return True

        except Exception as e:
            logger.error(f"Ошибка сохранения заказа: {e}")
            return False

    def get_student_orders(self, student_id: str, date_str: str) -> Dict[str, bool]:
        """Получает заказы ученика на дату"""
        try:
            wb = load_workbook(self.orders_path, data_only=True)
            ws = wb.active

            # Находим строку ученика
            student_row = None
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(r, 1).value) == student_id:
                    student_row = r
                    break

            if not student_row:
                return self._empty_meals()

            # Находим колонки для даты
            breakfast_col = None
            lunch_col = None
            snack_col = None

            for col in range(4, ws.max_column + 1):
                header = ws.cell(1, col).value
                if header and date_str in str(header):
                    if "_breakfast" in str(header):
                        breakfast_col = col
                    elif "_lunch" in str(header):
                        lunch_col = col
                    elif "_snack" in str(header):
                        snack_col = col

            if not all([breakfast_col, lunch_col, snack_col]):
                return self._empty_meals()

            # Получаем заказы
            orders = {
                'breakfast': ws.cell(row=student_row, column=breakfast_col).value == "✅",
                'lunch': ws.cell(row=student_row, column=lunch_col).value == "✅",
                'snack': ws.cell(row=student_row, column=snack_col).value == "✅"
            }

            return orders

        except Exception as e:
            logger.error(f"Ошибка получения заказов: {e}")
            return self._empty_meals()

    def _empty_meals(self) -> Dict[str, bool]:
        return {meal.value: False for meal in MealType}

    def count_for_date(self, date_str: str) -> Dict[str, int]:
        """Подсчет заказов на дату"""
        try:
            wb = load_workbook(self.orders_path, data_only=True)
            ws = wb.active

            # Находим колонки для даты
            breakfast_col = None
            lunch_col = None
            snack_col = None

            for col in range(4, ws.max_column + 1):
                header = ws.cell(1, col).value
                if header and date_str in str(header):
                    if "_breakfast" in str(header):
                        breakfast_col = col
                    elif "_lunch" in str(header):
                        lunch_col = col
                    elif "_snack" in str(header):
                        snack_col = col

            if not all([breakfast_col, lunch_col, snack_col]):
                return {meal.value: 0 for meal in MealType}

            # Подсчитываем
            counts = {meal.value: 0 for meal in MealType}
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, breakfast_col).value == "✅":
                    counts['breakfast'] += 1
                if ws.cell(row, lunch_col).value == "✅":
                    counts['lunch'] += 1
                if ws.cell(row, snack_col).value == "✅":
                    counts['snack'] += 1

            return counts

        except Exception as e:
            logger.error(f"Ошибка подсчета заказов: {e}")
            return {meal.value: 0 for meal in MealType}

    def get_working_dates(self, count: int = 10) -> List[Dict[str, str]]:
        """Получает список рабочих дат с проверкой блокировки"""
        dates = []
        today = get_current_datetime()
        added = 0
        current_date = today

        while added < count:
            if current_date.weekday() < 5:  # Только будни
                date_str = current_date.strftime("%Y-%m-%d")
                date_obj = current_date.date()

                dates.append({
                    'date_str': date_str,
                    'display': f"{current_date.strftime('%d.%m')} ({DAY_NAMES_RU[current_date.weekday()]})",
                    'is_locked': is_date_locked(date_obj)
                })
                added += 1
            current_date += timedelta(days=1)

        return dates

    def check_tomorrow_order(self, student_id: str) -> bool:
        """Проверяет, есть ли заказ на завтра"""
        tomorrow = (get_current_datetime() + timedelta(days=1)).strftime("%Y-%m-%d")
        orders = self.get_student_orders(student_id, tomorrow)

        # Проверяем, есть ли хотя бы один заказ
        return any(orders.values())

    def get_user_student_id(self, user_id: int, user_sessions: Dict) -> Optional[str]:
        """Получает ID ученика для пользователя"""
        if user_id in user_sessions and 'student_id' in user_sessions[user_id]:
            return user_sessions[user_id]['student_id']
        return None


# ================== КНОПКИ ==================
class KB:
    @staticmethod
    def main(has_reminder: bool = False):
        buttons = [
            [InlineKeyboardButton("🔑 Ввести ID ученика", callback_data="input_id")],
            [InlineKeyboardButton("📊 Статистика", callback_data="stats")],
            [
                InlineKeyboardButton(
                    f"{'🔔' if has_reminder else '🔕'} Напоминания: {'ВКЛ' if has_reminder else 'ВЫКЛ'}",
                    callback_data="toggle_reminder"
                )
            ]
        ]
        return InlineKeyboardMarkup(buttons)

    @staticmethod
    def dates(dates_list: List[Dict[str, str]]):
        keyboard = []
        for date_info in dates_list:
            display = date_info['display']
            if date_info['is_locked']:
                display = f"🔒 {display}"
            keyboard.append([
                InlineKeyboardButton(
                    display,
                    callback_data=f"date|{date_info['date_str']}"
                )
            ])
        keyboard.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_main")])
        return InlineKeyboardMarkup(keyboard)

    @staticmethod
    def meals(date_str: str, current_orders: Dict[str, bool]):
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        date_display = f"{date_obj.strftime('%d.%m.%Y')} ({DAY_NAMES_RU[date_obj.weekday()]})"

        # Проверяем блокировку
        is_locked = is_date_locked(date_obj.date())

        if is_locked:
            text = f"📅 {date_display}\n🔒 Редактирование закрыто (дедлайн: {Config.DEADLINE_TIME.strftime('%H:%M')})\n\nТекущий заказ:"
            buttons = [
                [InlineKeyboardButton(f"Завтрак: {'✅' if current_orders['breakfast'] else '❌'}",
                                      callback_data="locked")],
                [InlineKeyboardButton(f"Обед: {'✅' if current_orders['lunch'] else '❌'}", callback_data="locked")],
                [InlineKeyboardButton(f"Полдник: {'✅' if current_orders['snack'] else '❌'}", callback_data="locked")],
                [InlineKeyboardButton("⬅️ К датам", callback_data="back_dates")]
            ]
        else:
            text = f"📅 {date_display}\n✅ Можно редактировать (до {Config.DEADLINE_TIME.strftime('%H:%M')})\n\nВыберите питание:"
            buttons = [
                [
                    InlineKeyboardButton(
                        f"{'✅ ' if current_orders['breakfast'] else ''}Завтрак",
                        callback_data=f"meal|{date_str}|breakfast"
                    )
                ],
                [
                    InlineKeyboardButton(
                        f"{'✅ ' if current_orders['lunch'] else ''}Обед",
                        callback_data=f"meal|{date_str}|lunch"
                    )
                ],
                [
                    InlineKeyboardButton(
                        f"{'✅ ' if current_orders['snack'] else ''}Полдник",
                        callback_data=f"meal|{date_str}|snack"
                    )
                ],
                [
                    InlineKeyboardButton("✅ Всё на день", callback_data=f"all_day|{date_str}"),
                    InlineKeyboardButton("❌ Ничего", callback_data=f"none_day|{date_str}")
                ],
                [
                    InlineKeyboardButton("📅 Вся неделя", callback_data=f"all_week|{date_str}"),
                    InlineKeyboardButton("🗑️ Очистить неделю", callback_data=f"clear_week|{date_str}")
                ],
                [InlineKeyboardButton("⬅️ К датам", callback_data="back_dates")]
            ]

        return InlineKeyboardMarkup(buttons)

    @staticmethod
    def stats(is_admin: bool):
        buttons = []
        if is_admin:
            buttons.append([
                InlineKeyboardButton("📥 Скачать orders.xlsx", callback_data="download_orders")
            ])
            buttons.append([
                InlineKeyboardButton("📋 Скачать шаблон", callback_data="download_template")
            ])
            buttons.append([
                InlineKeyboardButton("🔄 Обновить данные", callback_data="refresh_data")
            ])
        buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_main")])
        return InlineKeyboardMarkup(buttons)


# ================== БОТ ==================
class FoodBot:
    INPUT_ID, DATES, MEALS = range(3)

    def __init__(self, application: Application):
        self.db = Database()
        self.user_sessions = {}
        self.application = application
        self._setup_reminder_job()

    def _setup_reminder_job(self):
        """Настраивает задачу напоминаний"""
        try:
            # Запускаем задачу напоминаний
            self.application.job_queue.run_repeating(
                self.send_reminders,
                interval=timedelta(minutes=1),  # Проверяем каждый час
                first=10  # Начинаем через 10 секунд после запуска
            )
            logger.info("Задача напоминаний настроена")
        except Exception as e:
            logger.error(f"Ошибка настройки задачи напоминаний: {e}")

    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик команды /start"""
        user_id = update.effective_user.id
        self.user_sessions[user_id] = {'state': 'main'}

        now = get_current_datetime()
        has_reminder = self.db.reminder_manager.get_user_reminder(user_id)

        await update.message.reply_text(
            f"🏫 **Система заказа школьного питания**\n\n"
            f"📅 Сегодня: {now.strftime('%d.%m.%Y')}\n"
            f"⏰ Напоминания: {'🔔 ВКЛЮЧЕНЫ (в 7:00)' if has_reminder else '🔕 ВЫКЛЮЧЕНЫ'}\n\n"
            f"Выберите действие:",
            parse_mode='Markdown',
            reply_markup=KB.main(has_reminder)
        )

    async def send_reminders(self, context: ContextTypes.DEFAULT_TYPE):
        """Отправляет напоминания о заказе еды"""
        try:
            now = get_current_datetime()
            current_time = now.time()

            # Проверяем, что сейчас 7:00 (или около того)
            if (
                    current_time.hour == Config.REMINDER_TIME.hour
                    and 0 <= current_time.minute <= 9
            ):
                logger.info(f"Проверка напоминаний в {current_time.strftime('%H:%M')}")

                # Получаем всех пользователей с включенными напоминаниями
                users_with_reminders = self.db.reminder_manager.get_all_users_with_reminders()

                for user_id in users_with_reminders:
                    try:
                        # Получаем ID ученика для пользователя
                        student_id = self.db.get_user_student_id_from_storage(user_id)


                        if student_id:
                            # Проверяем, есть ли заказ на завтра
                            has_order = self.db.check_tomorrow_order(student_id)

                            if not has_order:
                                # Получаем информацию об ученике
                                ok, student_info = self.db.verify_student(student_id)

                                if ok:
                                    tomorrow = (now + timedelta(days=1)).strftime("%d.%m.%Y")
                                    message = (
                                        f"🔔 **Напоминание о заказе питания**\n\n"
                                        f"👤 {student_info.full_name}\n"
                                        f"🏫 {student_info.class_name}\n\n"
                                        f"📅 **На завтра ({tomorrow}) у вас нет заказа!**\n\n"
                                        f"⏰ Дедлайн заказа: {Config.DEADLINE_TIME.strftime('%H:%M')}\n"
                                        f"⚡ Успейте сделать заказ до дедлайна!"
                                    )

                                    await context.bot.send_message(
                                        chat_id=user_id,
                                        text=message,
                                        parse_mode='Markdown'
                                    )
                                    logger.info(f"Отправлено напоминание пользователю {user_id}")
                        else:
                            # Если у пользователя нет активной сессии с учеником
                            logger.debug(f"У пользователя {user_id} нет активной сессии с учеником")

                    except Exception as e:
                        logger.error(f"Ошибка отправки напоминания пользователю {user_id}: {e}")

                logger.info(f"Напоминания отправлены для {len(users_with_reminders)} пользователей")

        except Exception as e:
            logger.error(f"Ошибка в задаче напоминаний: {e}")

    async def button_handler(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик нажатий на кнопки"""
        query = update.callback_query
        await query.answer()

        user_id = query.from_user.id
        data = query.data

        if data == "input_id":
            await query.edit_message_text(
                "🔑 **Введите ID ученика**\n\n"
                "ID можно получить у классного руководителя.\n"
                "**Введите ID:**",
                parse_mode='Markdown'
            )
            return self.INPUT_ID

        elif data == "toggle_reminder":
            # Переключаем напоминание
            new_state = self.db.reminder_manager.toggle_user_reminder(user_id)

            now = get_current_datetime()
            await query.edit_message_text(
                f"🏫 **Система заказа школьного питания**\n\n"
                f"📅 Сегодня: {now.strftime('%d.%m.%Y')}\n"
                f"⏰ Напоминания: {'🔔 ВКЛЮЧЕНЫ (в 7:00)' if new_state else '🔕 ВЫКЛЮЧЕНЫ'}\n\n"
                f"{'✅ Напоминания включены! Буду напоминать в 7:00 утра.' if new_state else '❌ Напоминания отключены.'}\n\n"
                f"Выберите действие:",
                parse_mode='Markdown',
                reply_markup=KB.main(new_state)
            )
            return

        elif data == "stats":
            if user_id not in Config.ADMIN_IDS:
                await query.edit_message_text(
                    "❌ У вас нет доступа к статистике",
                    reply_markup=KB.main(self.db.reminder_manager.get_user_reminder(user_id))
                )
                return

            # Получаем статистику
            today = get_current_datetime().strftime("%Y-%m-%d")
            tomorrow = (get_current_datetime() + timedelta(days=1)).strftime("%Y-%m-%d")

            today_stats = self.db.count_for_date(today)
            tomorrow_stats = self.db.count_for_date(tomorrow)

            text = (
                "📊 **Статистика заказов**\n\n"
                f"**Сегодня ({get_current_datetime().strftime('%d.%m')}):**\n"
                f"🍳 Завтрак: {today_stats['breakfast']}\n"
                f"🍲 Обед: {today_stats['lunch']}\n"
                f"🥪 Полдник: {today_stats['snack']}\n\n"
                f"**Завтра ({datetime.fromisoformat(tomorrow).strftime('%d.%m')}):**\n"
                f"🍳 Завтрак: {tomorrow_stats['breakfast']}\n"
                f"🍲 Обед: {tomorrow_stats['lunch']}\n"
                f"🥪 Полдник: {tomorrow_stats['snack']}"
            )

            await query.edit_message_text(
                text,
                parse_mode='Markdown',
                reply_markup=KB.stats(is_admin=True)
            )

        elif data == "download_orders":
            if user_id not in Config.ADMIN_IDS:
                return

            if os.path.exists(self.db.orders_path):
                await query.message.reply_document(
                    document=open(self.db.orders_path, 'rb'),
                    filename="orders.xlsx",
                    caption="📊 Файл заказов"
                )

        elif data == "download_template":
            if user_id not in Config.ADMIN_IDS:
                return

            if os.path.exists(self.db.template_path):
                await query.message.reply_document(
                    document=open(self.db.template_path, 'rb'),
                    filename=Config.TEMPLATE_FILE,
                    caption="📋 Основной шаблон"
                )

        elif data == "refresh_data":
            if user_id not in Config.ADMIN_IDS:
                return

            # Перезагружаем шаблон
            if self.db.template_manager.load_template():
                await self._send_temp_message(
                    query.message.chat_id,
                    "✅ Данные обновлены",
                    context
                )
            else:
                await self._send_temp_message(
                    query.message.chat_id,
                    "❌ Ошибка обновления данных",
                    context
                )

        elif data == "back_main":
            if user_id in self.user_sessions:
                self.user_sessions[user_id] = {'state': 'main'}

            now = get_current_datetime()
            has_reminder = self.db.reminder_manager.get_user_reminder(user_id)

            await query.edit_message_text(
                f"🏫 **Система заказа школьного питания**\n\n"
                f"📅 Сегодня: {now.strftime('%d.%m.%Y')}\n"
                f"⏰ Напоминания: {'🔔 ВКЛЮЧЕНЫ (в 7:00)' if has_reminder else '🔕 ВЫКЛЮЧЕНЫ'}\n\n"
                f"Выберите действие:",
                parse_mode='Markdown',
                reply_markup=KB.main(has_reminder)
            )

        elif data == "back_dates":
            if user_id not in self.user_sessions or 'student_id' not in self.user_sessions[user_id]:
                await query.edit_message_text(
                    "❌ Сессия устарела. Начните заново.",
                    reply_markup=KB.main(self.db.reminder_manager.get_user_reminder(user_id))
                )
                return

            dates = self.db.get_working_dates(10)
            student_info = self.user_sessions[user_id]

            await query.edit_message_text(
                f"👤 **{student_info['student_name']}**\n"
                f"🏫 Класс: {student_info['class_name']}\n\n"
                f"Выберите дату (🔒 - редактирование закрыто):",
                parse_mode='Markdown',
                reply_markup=KB.dates(dates)
            )
            return self.DATES

        elif data.startswith("date|"):
            date_str = data.split("|")[1]

            if user_id not in self.user_sessions or 'student_id' not in self.user_sessions[user_id]:
                await query.edit_message_text(
                    "❌ Сессия устарела. Начните заново.",
                    reply_markup=KB.main(self.db.reminder_manager.get_user_reminder(user_id))
                )
                return

            student_info = self.user_sessions[user_id]
            orders = self.db.get_student_orders(student_info['student_id'], date_str)

            await query.edit_message_text(
                f"📅 **{datetime.strptime(date_str, '%Y-%m-%d').strftime('%d.%m.%Y')}**\n"
                f"👤 {student_info['student_name']}\n"
                f"🏫 {student_info['class_name']}",
                parse_mode='Markdown',
                reply_markup=KB.meals(date_str, orders)
            )
            return self.MEALS

        elif data == "locked":
            await self._send_temp_message(
                query.message.chat_id,
                "⛔ Редактирование заказов на эту дату закрыто",
                context
            )
            return self.MEALS

        elif data.startswith("meal|"):
            _, date_str, meal_type = data.split("|")

            if user_id not in self.user_sessions or 'student_id' not in self.user_sessions[user_id]:
                return

            student_info = self.user_sessions[user_id]

            # Проверяем можно ли редактировать
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            if is_date_locked(target_date):
                await self._send_temp_message(
                    query.message.chat_id,
                    f"⛔ Редактирование заказов на эту дату закрыто (дедлайн: {Config.DEADLINE_TIME.strftime('%H:%M')})",
                    context
                )
                return self.MEALS

            # Получаем и обновляем заказы
            orders = self.db.get_student_orders(student_info['student_id'], date_str)
            orders[meal_type] = not orders[meal_type]

            # Сохраняем
            if self.db.save_order(student_info['student_id'], date_str, orders):
                await query.edit_message_reply_markup(
                    KB.meals(date_str, orders)
                )
                await self._send_temp_message(
                    query.message.chat_id,
                    "✅ Заказ обновлен",
                    context
                )
            else:
                await self._send_temp_message(
                    query.message.chat_id,
                    "❌ Ошибка сохранения заказа",
                    context
                )

        elif data.startswith("all_day|"):
            date_str = data.split("|")[1]

            if user_id not in self.user_sessions or 'student_id' not in self.user_sessions[user_id]:
                return

            # Проверяем можно ли редактировать
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            if is_date_locked(target_date):
                await self._send_temp_message(
                    query.message.chat_id,
                    f"⛔ Редактирование заказов на эту дату закрыто (дедлайн: {Config.DEADLINE_TIME.strftime('%H:%M')})",
                    context
                )
                return self.MEALS

            # Заказываем всё на день
            orders = {meal.value: True for meal in MealType}

            if self.db.save_order(self.user_sessions[user_id]['student_id'], date_str, orders):
                await query.edit_message_reply_markup(
                    KB.meals(date_str, orders)
                )
                await self._send_temp_message(
                    query.message.chat_id,
                    "✅ Заказано всё питание на день",
                    context
                )

        elif data.startswith("none_day|"):
            date_str = data.split("|")[1]

            if user_id not in self.user_sessions or 'student_id' not in self.user_sessions[user_id]:
                return

            # Проверяем можно ли редактировать
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            if is_date_locked(target_date):
                await self._send_temp_message(
                    query.message.chat_id,
                    f"⛔ Редактирование заказов на эту дату закрыто (дедлайн: {Config.DEADLINE_TIME.strftime('%H:%M')})",
                    context
                )
                return self.MEALS

            # Отменяем всё на день
            orders = {meal.value: False for meal in MealType}

            if self.db.save_order(self.user_sessions[user_id]['student_id'], date_str, orders):
                await query.edit_message_reply_markup(
                    KB.meals(date_str, orders)
                )
                await self._send_temp_message(
                    query.message.chat_id,
                    "❌ Питание на день отменено",
                    context
                )

        elif data.startswith("all_week|"):
            date_str = data.split("|")[1]

            if user_id not in self.user_sessions or 'student_id' not in self.user_sessions[user_id]:
                return

            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            monday = date_obj - timedelta(days=date_obj.weekday())

            success = 0
            total = 0

            for i in range(5):  # Понедельник - Пятница
                week_date = monday + timedelta(days=i)
                week_date_str = week_date.strftime("%Y-%m-%d")

                # Пропускаем заблокированные даты
                if is_date_locked(week_date.date()):
                    continue

                total += 1
                orders = {meal.value: True for meal in MealType}

                if self.db.save_order(self.user_sessions[user_id]['student_id'], week_date_str, orders):
                    success += 1

            if success > 0:
                await self._send_temp_message(
                    query.message.chat_id,
                    f"✅ Заказано питание на {success} дней недели",
                    context
                )

            # Обновляем текущий день
            current_orders = self.db.get_student_orders(
                self.user_sessions[user_id]['student_id'], date_str
            )
            await query.edit_message_reply_markup(
                KB.meals(date_str, current_orders)
            )

        elif data.startswith("clear_week|"):
            date_str = data.split("|")[1]

            if user_id not in self.user_sessions or 'student_id' not in self.user_sessions[user_id]:
                return

            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            monday = date_obj - timedelta(days=date_obj.weekday())

            success = 0
            total = 0

            for i in range(5):  # Понедельник - Пятница
                week_date = monday + timedelta(days=i)
                week_date_str = week_date.strftime("%Y-%m-%d")

                # Пропускаем заблокированные даты
                if is_date_locked(week_date.date()):
                    continue

                total += 1
                orders = {meal.value: False for meal in MealType}

                if self.db.save_order(self.user_sessions[user_id]['student_id'], week_date_str, orders):
                    success += 1

            if success > 0:
                await self._send_temp_message(
                    query.message.chat_id,
                    f"❌ Питание отменено на {success} дней недели",
                    context
                )

            # Обновляем текущий день
            current_orders = self.db.get_student_orders(
                self.user_sessions[user_id]['student_id'], date_str
            )
            await query.edit_message_reply_markup(
                KB.meals(date_str, current_orders)
            )

    async def input_id_handler(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик ввода ID ученика"""
        user_id = update.effective_user.id
        student_id = update.message.text.strip()

        # Проверяем ID (должен быть числовым)
        if not student_id.isdigit():
            await update.message.reply_text(
                "❌ **Неверный формат ID**\n\n"
                "ID должен состоять только из цифр.\n"
                "Пример правильного ID: 100953\n\n"
                "**Попробуйте снова:**",
                parse_mode='Markdown'
            )
            return self.INPUT_ID

        # Проверяем ID
        ok, student_info = self.db.verify_student(student_id)

        if not ok:
            await update.message.reply_text(
                "❌ **Ученик с таким ID не найден**\n\n"
                "Пожалуйста, проверьте ID и попробуйте снова.\n"
                "ID можно получить у классного руководителя.\n\n"
                "**Введите ID еще раз:**",
                parse_mode='Markdown'
            )
            return self.INPUT_ID

        # Сохраняем в сессию
        self.user_sessions[user_id] = {
            'student_id': student_id,
            'student_name': student_info.full_name,
            'class_name': student_info.class_name,
            'state': 'dates'
        }

        # Показываем доступные даты
        dates = self.db.get_working_dates(10)

        await update.message.reply_text(
            f"✅ **Ученик найден!**\n\n"
            f"👤 **{student_info.full_name}**\n"
            f"🏫 Класс: {student_info.class_name}\n"
            f"🔑 ID: {student_id}\n\n"
            f"Выберите дату (🔒 - редактирование закрыто):",
            parse_mode='Markdown',
            reply_markup=KB.dates(dates)
        )

        return self.DATES

    async def cancel(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Обработчик отмены"""
        user_id = update.effective_user.id
        if user_id in self.user_sessions:
            self.user_sessions[user_id] = {'state': 'main'}

        has_reminder = self.db.reminder_manager.get_user_reminder(user_id)
        await update.message.reply_text(
            "❌ Действие отменено",
            reply_markup=KB.main(has_reminder)
        )
        return ConversationHandler.END

    async def _send_temp_message(self, chat_id: int, text: str, context: ContextTypes.DEFAULT_TYPE, delay: int = 2):
        """Отправляет временное сообщение"""
        msg = await context.bot.send_message(chat_id=chat_id, text=text)
        await asyncio.sleep(delay)
        try:
            await msg.delete()
        except:
            pass

    async def time_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Команда для проверки времени"""
        now = get_current_datetime()
        deadline_time = Config.DEADLINE_TIME
        reminder_time = Config.REMINDER_TIME

        message = (
            f"🕐 **Текущее время:** {now.strftime('%H:%M:%S')}\n"
            f"📅 **Дата:** {now.strftime('%d.%m.%Y')}\n"
            f"⏰ **Дедлайн редактирования:** {deadline_time.strftime('%H:%M')}\n"
            f"🔔 **Время напоминаний:** {reminder_time.strftime('%H:%M')}\n"
            f"🔒 **Сегодняшний день заблокирован:** {'Да' if now.time() >= deadline_time else 'Нет'}\n"
        )

        await update.message.reply_text(message, parse_mode='Markdown')

    async def test_deadline(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Тест дедлайна (только для админов)"""
        if update.effective_user.id not in Config.ADMIN_IDS:
            return

        now = get_current_datetime()
        test_dates = [
            (now.date(), "Сегодня"),
            (now.date() + timedelta(days=1), "Завтра"),
            (now.date() - timedelta(days=1), "Вчера"),
        ]

        results = []
        for test_date, name in test_dates:
            locked = is_date_locked(test_date)
            results.append(f"{name} ({test_date}): {'🔒 ЗАБЛОКИРОВАНО' if locked else '✅ ДОСТУПНО'}")

        await update.message.reply_text(
            "🧪 **Тест дедлайна**\n\n" + "\n".join(results),
            parse_mode='Markdown'
        )

    async def reminder_info(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Информация о напоминаниях"""
        user_id = update.effective_user.id
        has_reminder = self.db.reminder_manager.get_user_reminder(user_id)

        message = (
            f"🔔 **Информация о напоминаниях**\n\n"
            f"📅 **Статус:** {'🔔 ВКЛЮЧЕНЫ' if has_reminder else '🔕 ВЫКЛЮЧЕНЫ'}\n"
            f"⏰ **Время напоминания:** {Config.REMINDER_TIME.strftime('%H:%M')}\n"
            f"⏳ **Что проверяется:** Заказ на завтрашний день\n"
            f"📝 **Условие:** Напоминание приходит, если на завтра нет ни одного заказа\n\n"
            f"ℹ️ Напоминания можно включить/выключить в главном меню"
        )

        await update.message.reply_text(message, parse_mode='Markdown')


# ================== ЗАПУСК ==================
def main():
    """Основная функция запуска бота"""
    if not Config.BOT_TOKEN:
        logger.error("❌ Не указан BOT_TOKEN в конфигурации!")
        print("=" * 50)
        print("ВНИМАНИЕ: Не указан токен бота!")
        print("Добавьте в код строку: Config.BOT_TOKEN = 'ВАШ_ТОКЕН'")
        print("=" * 50)
        return

    # Создаем приложение С JobQueue
    application = (
        Application.builder()
        .token(Config.BOT_TOKEN)
        .build()
    )

    # Создаем бота и передаем ему application
    bot = FoodBot(application)

    # Добавляем обработчики команд
    application.add_handler(CommandHandler("start", bot.start))
    application.add_handler(CommandHandler("cancel", bot.cancel))
    application.add_handler(CommandHandler("time", bot.time_command))
    application.add_handler(CommandHandler("test", bot.test_deadline))
    application.add_handler(CommandHandler("reminder", bot.reminder_info))

    # Добавляем ConversationHandler для ввода ID
    conv_handler = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(bot.button_handler, pattern="^input_id$")
        ],
        states={
            bot.INPUT_ID: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, bot.input_id_handler)
            ],
            bot.DATES: [
                CallbackQueryHandler(bot.button_handler)
            ],
            bot.MEALS: [
                CallbackQueryHandler(bot.button_handler)
            ]
        },
        fallbacks=[
            CommandHandler("cancel", bot.cancel),
            CallbackQueryHandler(bot.button_handler, pattern="^back_main$")
        ],
        allow_reentry=True
    )

    application.add_handler(conv_handler)

    # Обработчик остальных кнопок
    application.add_handler(CallbackQueryHandler(bot.button_handler))

    # Запускаем бота
    logger.info("🤖 Бот запускается...")

    print("\n" + "=" * 50)
    print("🏫 Школьный бот питания")
    print("=" * 50)
    print(f"Текущее время: {get_current_datetime().strftime('%H:%M:%S')}")
    print(f"Дедлайн редактирования: {Config.DEADLINE_TIME.strftime('%H:%M')}")
    print(f"Время напоминаний: {Config.REMINDER_TIME.strftime('%H:%M')}")
    print(f"Часовой пояс: UTC+{Config.TIMEZONE_OFFSET}")
    print("=" * 50)

    # Проверяем файлы
    required_files = [
        (bot.db.students_path, "students.xlsx"),
        (bot.db.template_path, "шаблон.xlsx")
    ]

    for file_path, name in required_files:
        if os.path.exists(file_path):
            print(f"✅ {name}: найден")
        else:
            print(f"⚠️  {name}: не найден")

    print("=" * 50)
    print("Команды для проверки:")
    print("/time - текущее время и статус дедлайна")
    print("/test - тест дедлайна (только для админов)")
    print("/reminder - информация о напоминаниях")
    print("=" * 50 + "\n")

    try:
        application.run_polling(allowed_updates=Update.ALL_TYPES)
    except Exception as e:
        logger.error(f"Ошибка запуска бота: {e}")
        print(f"\n❌ Ошибка: {e}")


if __name__ == "__main__":
    main()






